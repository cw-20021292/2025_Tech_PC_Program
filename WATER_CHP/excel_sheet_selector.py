"""
Excel 파일 읽기 및 Sheet 선택 모듈
Excel 파일을 업로드하고 Sheet를 선택하는 기능을 제공합니다.
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import openpyxl


class ExcelSheetSelector:
    """Excel 파일 읽기 및 Sheet 선택 클래스"""
    
    def __init__(self, parent_window):
        """
        Args:
            parent_window: 부모 윈도우 (Tkinter root)
        """
        self.parent_window = parent_window
        self.selected_file_path = None
        self.selected_sheet_name = None
        self.workbook = None
    
    def select_excel_file(self):
        """
        Excel 파일 선택 다이얼로그를 열고 파일을 선택합니다.
        
        Returns:
            str: 선택된 파일 경로, 취소 시 None
        """
        file_path = filedialog.askopenfilename(
            title="Excel 파일 선택",
            filetypes=[
                ("Excel 파일", "*.xlsx *.xls"),
                ("모든 파일", "*.*")
            ]
        )
        
        if not file_path:
            return None
        
        try:
            # Excel 파일 열기 (data_only=True로 수식이 아닌 값만 읽기)
            self.workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            self.selected_file_path = file_path
            return file_path
        except Exception as e:
            messagebox.showerror(
                "오류",
                f"Excel 파일을 읽을 수 없습니다.\n{str(e)}"
            )
            return None
    
    def get_sheet_names(self):
        """
        현재 열려있는 Excel 파일의 Sheet 이름 목록을 반환합니다.
        
        Returns:
            list: Sheet 이름 목록, 파일이 열려있지 않으면 None
        """
        if self.workbook is None:
            return None
        
        return self.workbook.sheetnames
    
    def show_sheet_selection_dialog(self, callback=None):
        """
        Sheet 선택 다이얼로그를 표시합니다.
        
        Args:
            callback: Sheet 선택 시 호출될 콜백 함수 (selected_sheet_name)
        
        Returns:
            str: 선택된 Sheet 이름, 취소 시 None
        """
        # 먼저 Excel 파일 선택
        file_path = self.select_excel_file()
        if file_path is None:
            return None
        
        # Sheet 이름 목록 가져오기
        sheet_names = self.get_sheet_names()
        if not sheet_names:
            messagebox.showwarning(
                "경고",
                "Excel 파일에 Sheet가 없습니다."
            )
            return None
        
        # Sheet 선택 다이얼로그 생성
        dialog = tk.Toplevel(self.parent_window)
        dialog.title("Sheet 선택")
        dialog.geometry("400x400")
        dialog.transient(self.parent_window)
        dialog.grab_set()
        
        # 중앙 정렬을 위한 설정
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # 다이얼로그를 grid로 구성하여 버튼 영역 고정
        dialog.columnconfigure(0, weight=1)
        dialog.rowconfigure(1, weight=1)  # 스크롤 영역만 확장
        
        selected_sheet = tk.StringVar()
        selected_sheet.set(sheet_names[0] if sheet_names else "")
        
        # 제목 라벨
        file_name = os.path.basename(file_path)
        title_label = ttk.Label(
            dialog,
            text=f"Excel 파일: {file_name}\n\nSheet를 선택하세요:",
            font=("Arial", 10)
        )
        title_label.grid(row=0, column=0, pady=10, sticky=(tk.W, tk.E))
        
        # 스크롤 가능한 Sheet 목록 프레임
        scroll_container = ttk.Frame(dialog)
        scroll_container.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=20, pady=10)
        
        # Canvas와 Scrollbar 생성
        canvas = tk.Canvas(scroll_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        # 스크롤 가능한 프레임을 Canvas에 배치
        def on_frame_configure(event):
            """내부 프레임 크기 변경 시 스크롤 영역 업데이트"""
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        scrollable_frame.bind("<Configure>", on_frame_configure)
        
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Canvas 크기 변경 시 내부 프레임 너비 조정
        def on_canvas_configure(event):
            canvas_width = event.width
            canvas.itemconfig(canvas_window, width=canvas_width)
        
        canvas.bind("<Configure>", on_canvas_configure)
        
        # Canvas와 Scrollbar 배치
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Sheet 목록 (라디오 버튼) - 스크롤 가능한 프레임에 배치
        for i, sheet_name in enumerate(sheet_names):
            rb = ttk.Radiobutton(
                scrollable_frame,
                text=sheet_name,
                variable=selected_sheet,
                value=sheet_name
            )
            rb.pack(anchor=tk.W, pady=2)
        
        # 마우스 휠 이벤트 바인딩
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        def bind_to_mousewheel(event):
            canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        def unbind_from_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind("<Enter>", bind_to_mousewheel)
        canvas.bind("<Leave>", unbind_from_mousewheel)
        
        # 버튼 프레임 (항상 하단에 고정)
        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=2, column=0, pady=10, sticky=(tk.W, tk.E))
        
        def on_ok():
            """확인 버튼 클릭 시"""
            sheet_name = selected_sheet.get()
            if sheet_name:
                self.selected_sheet_name = sheet_name
                dialog.destroy()
                if callback:
                    callback(sheet_name)
            else:
                messagebox.showwarning("경고", "Sheet를 선택해주세요.")
        
        def on_cancel():
            """취소 버튼 클릭 시"""
            self.selected_sheet_name = None
            dialog.destroy()
        
        # 확인/취소 버튼
        ok_button = ttk.Button(button_frame, text="확인", command=on_ok)
        ok_button.pack(side=tk.LEFT, padx=5)
        
        cancel_button = ttk.Button(button_frame, text="취소", command=on_cancel)
        cancel_button.pack(side=tk.LEFT, padx=5)
        
        # 다이얼로그가 닫힐 때까지 대기
        dialog.wait_window()
        
        return self.selected_sheet_name
    
    def get_selected_sheet(self):
        """
        선택된 Sheet 이름을 반환합니다.
        
        Returns:
            str: 선택된 Sheet 이름, 없으면 None
        """
        return self.selected_sheet_name
    
    def get_workbook(self):
        """
        현재 열려있는 Workbook 객체를 반환합니다.
        
        Returns:
            openpyxl.workbook.Workbook: Workbook 객체, 없으면 None
        """
        return self.workbook
    
    def read_icemaking_table_data(self, sheet_name):
        """
        제빙테이블 데이터를 읽어옵니다.
        
        데이터 구조:
        - A4~A49: 입수온도 (세로 헤더, 46개)
        - B3~AU3: 외기온도 (가로 헤더, 46개)
        - B4~AU49: 테이블 데이터 (46행 x 46열)
        
        Args:
            sheet_name: 읽을 Sheet 이름
            
        Returns:
            dict: {
                'water_temps': [A4~A49 값들],
                'outdoor_temps': [B3~AU3 값들],
                'table_data': [[B4~AU4], [B5~AU5], ..., [B49~AU49]]
            }
            실패 시 None
        """
        if self.workbook is None:
            return None
        
        try:
            # Sheet 선택
            sheet = self.workbook[sheet_name]
            
            # 입수온도 읽기 (A4~A49, 46개)
            water_temps = []
            for row in range(4, 50):  # A4~A49
                cell_value = sheet[f'A{row}'].value
                if cell_value is not None:
                    water_temps.append(float(cell_value))
                else:
                    water_temps.append(0.0)
            
            # 외기온도 읽기 (B3~AU3, 46개)
            outdoor_temps = []
            # B=2열부터 AU=47열까지
            for col in range(2, 48):  # B(2) ~ AU(47)
                cell_value = sheet.cell(row=3, column=col).value
                if cell_value is not None:
                    outdoor_temps.append(float(cell_value))
                else:
                    outdoor_temps.append(0.0)
            
            # 테이블 데이터 읽기 (B4~AU49, 46행 x 46열)
            table_data = []
            for row in range(4, 50):  # 4~49행
                row_data = []
                for col in range(2, 48):  # B(2)~AU(47)열
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(float(cell_value))
                    else:
                        row_data.append(0.0)
                table_data.append(row_data)
            
            return {
                'water_temps': water_temps,
                'outdoor_temps': outdoor_temps,
                'table_data': table_data
            }
            
        except Exception as e:
            messagebox.showerror(
                "오류",
                f"제빙테이블 데이터를 읽을 수 없습니다.\n{str(e)}"
            )
            return None
    
    def read_sheet_data(self, sheet_name):
        """
        선택된 Sheet의 데이터를 읽어옵니다.
        
        Args:
            sheet_name: 읽을 Sheet 이름
            
        Returns:
            dict: 읽은 데이터 {'outdoor_temps': [], 'water_temps': [], 'table_data': []}
                  실패 시 None
        """
        if not self.workbook or not sheet_name:
            return None
        
        try:
            sheet = self.workbook[sheet_name]
            
            # 외기온도 헤더 읽기 (B3~AU3, 열 인덱스 2~47)
            outdoor_temps = []
            for col in range(2, 48):  # B(2)부터 AU(47)까지 46개
                cell_value = sheet.cell(row=3, column=col).value
                if cell_value is not None:
                    try:
                        outdoor_temps.append(int(float(cell_value)))
                    except (ValueError, TypeError):
                        outdoor_temps.append(0)
                else:
                    outdoor_temps.append(0)
            
            # 입수온도 헤더 읽기 (A4~A49, 행 인덱스 4~49)
            water_temps = []
            for row in range(4, 50):  # A4(4)부터 A49(49)까지 46개
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value is not None:
                    try:
                        water_temps.append(int(float(cell_value)))
                    except (ValueError, TypeError):
                        water_temps.append(0)
                else:
                    water_temps.append(0)
            
            # 테이블 데이터 읽기 (B4~AU49)
            table_data = []
            for row in range(4, 50):  # 행 4~49 (46개 행)
                row_data = []
                for col in range(2, 48):  # 열 B~AU (46개 열)
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        try:
                            row_data.append(int(float(cell_value)))
                        except (ValueError, TypeError):
                            row_data.append(0)
                    else:
                        row_data.append(0)
                table_data.append(row_data)
            
            return {
                'outdoor_temps': outdoor_temps,
                'water_temps': water_temps,
                'table_data': table_data
            }
        
        except Exception as e:
            print(f"Sheet 데이터 읽기 오류: {str(e)}")
            return None
    
    def close_workbook(self):
        """열려있는 Workbook을 닫습니다."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.selected_file_path = None
            self.selected_sheet_name = None

